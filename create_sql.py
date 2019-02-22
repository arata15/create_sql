import pandas as pd
import openpyxl as px
import re
import json
import requests as rq

#行の値,ファイルの番号
count = file_count = 1
#ループの条件
end_flg = False
#EXCELファイル内のワークブック読み込み
work_book = px.load_workbook("任意のディレクトリ/Excelファイル名")
#シートの情報読み込み
sheet = work_book.active
sheet = work_book.get_sheet_by_name("Excelシート名")
#SQLに設定する値
text = file_name = ""
upd_tbl_nm = sheet.cell(row=2, column=4)
upd_column = sheet.cell(row=2, column=5)
where_column = sheet.cell(row=2, column=6)

#SQL文の記載されたファイル作成
def create_file(update_text,where_text,text,file_name,file_count):
    file_name = "任意のディレクトリ/SQL" + str(file_count) + ".txt"
    text = update_text + "\n" + "END" + "\n" + where_text[:-1] + ")"
    file = open(file_name, "w")
    file.write(text)
    file.close()

#SQL文作成
try:
    while end_flg == False:
        count += 1
        columnA = sheet.cell(row=count, column=1)
        columnB = sheet.cell(row=count, column=2)
        if count % 1000 == 2 or count == 2:
            update_text = "UPDATE " + str(upd_tbl_nm.value) + " SET " + str(upd_column.value) + " = CASE " + str(where_column.value)
            where_text = "WHERE " + str(where_column.value) + " IN("
        
        if str(columnA.value) == str(None):
            create_file(update_text,where_text,text,file_name,file_count)
            end_flg = True
        else:
            update_text = update_text + "\nWHEN" + "'" + str(columnA.value) + "'"  + "THEN" + "'" + str(columnB.value) + "'"
            where_text = where_text + "\n'" + str(columnA.value) + "'," 
            if count % 1000 == 1:
                create_file(update_text,where_text,text,file_name,file_count)
                file_count += 1
except:
    import traceback
    traceback.print_exc()


